import os
import sys
import socket
from datetime import datetime
from functools import wraps

from flask import (Flask, render_template, request, redirect,
                   url_for, session, flash, g, send_from_directory)
from werkzeug.security import generate_password_hash, check_password_hash
from sqlalchemy import create_engine, text

# ---------------------------------------------------------------------------
# PyInstaller compatibility
# ---------------------------------------------------------------------------
if getattr(sys, "frozen", False):
    BASE_DIR = sys._MEIPASS
    EXE_DIR  = os.path.dirname(sys.executable)
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
    EXE_DIR  = BASE_DIR

app = Flask(
    __name__,
    template_folder=os.path.join(BASE_DIR, "templates"),
    static_folder=os.path.join(BASE_DIR, "static"),
)
app.secret_key = os.environ.get("SECRET_KEY", "banik-hardware-secret-2024")

# ---------------------------------------------------------------------------
# Database — PostgreSQL (cloud) or SQLite (local / EXE)
# ---------------------------------------------------------------------------
DB_PATH  = os.path.join(EXE_DIR, "inventory.db")
_db_url  = os.environ.get("DATABASE_URL", "").strip()

# Convert postgres:// or postgresql:// to postgresql+pg8000:// for pure-Python DB driver
if _db_url.startswith("postgres://"):
    _db_url = _db_url.replace("postgres://", "postgresql+pg8000://", 1)
elif _db_url.startswith("postgresql://") and "+pg8000" not in _db_url and "+psycopg" not in _db_url:
    _db_url = _db_url.replace("postgresql://", "postgresql+pg8000://", 1)

IS_POSTGRES = bool(_db_url)
_engine = create_engine(
    _db_url if IS_POSTGRES else f"sqlite:///{DB_PATH}",
    pool_pre_ping=True,
)

ORDER_BY = "LOWER(material)" if IS_POSTGRES else "material COLLATE NOCASE"


# ---------------------------------------------------------------------------
# Database helpers
# ---------------------------------------------------------------------------
def get_db():
    if "db" not in g:
        g.db = _engine.connect()
    return g.db


@app.teardown_appcontext
def close_db(exception=None):
    db = g.pop("db", None)
    if db is not None:
        db.close()


def init_db():
    """Create tables and seed default users if not present. Safe to call multiple times."""
    id_col = "id SERIAL PRIMARY KEY" if IS_POSTGRES else "id INTEGER PRIMARY KEY AUTOINCREMENT"
    with _engine.connect() as conn:
        conn.execute(text("""
            CREATE TABLE IF NOT EXISTS users (
                username      TEXT PRIMARY KEY,
                password_hash TEXT NOT NULL,
                role          TEXT NOT NULL CHECK (role IN ('admin', 'user'))
            )
        """))
        conn.execute(text(f"""
            CREATE TABLE IF NOT EXISTS inventory (
                {id_col},
                material        TEXT NOT NULL,
                buying_price    REAL NOT NULL,
                wholesale_price REAL NOT NULL,
                retail_price    REAL NOT NULL,
                updated_at      TEXT NOT NULL
            )
        """))
        conn.execute(text("""
            CREATE TABLE IF NOT EXISTS vendor_payments (
                inventory_id  INTEGER PRIMARY KEY,
                vendor_name   TEXT    NOT NULL DEFAULT '',
                vendor_phone  TEXT    NOT NULL DEFAULT '',
                total_amount  REAL    NOT NULL DEFAULT 0,
                paid_amount   REAL    NOT NULL DEFAULT 0,
                updated_at    TEXT    NOT NULL
            )
        """))
        row = conn.execute(text("SELECT username FROM users LIMIT 1")).fetchone()
        if row is None:
            conn.execute(
                text("INSERT INTO users (username, password_hash, role) VALUES (:u, :p, :r)"),
                [
                    {"u": "admin", "p": generate_password_hash("admin123"), "r": "admin"},
                    {"u": "user",  "p": generate_password_hash("user123"),  "r": "user"},
                ],
            )
        conn.commit()


_db_initialized = False

def ensure_db():
    global _db_initialized
    if not _db_initialized:
        try:
            init_db()
            _db_initialized = True
        except Exception as err:
            print(f"Database init retry: {err}")

@app.before_request
def auto_init_db():
    ensure_db()

# Safe initial call on startup
try:
    init_db()
    _db_initialized = True
except Exception as err:
    print(f"Startup DB init deferred: {err}")



# ---------------------------------------------------------------------------
# Auth helpers
# ---------------------------------------------------------------------------
def login_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if "username" not in session:
            return redirect(url_for("login"))
        return view(*args, **kwargs)
    return wrapped


def admin_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if session.get("role") != "admin":
            flash("You don't have permission to perform that action.", "error")
            return redirect(url_for("dashboard"))
        return view(*args, **kwargs)
    return wrapped


# ---------------------------------------------------------------------------
# Routes — PWA static files
# ---------------------------------------------------------------------------
@app.route("/manifest.json")
def manifest():
    return send_from_directory(app.static_folder, "manifest.json",
                               mimetype="application/manifest+json")


@app.route("/sw.js")
def service_worker():
    return send_from_directory(app.static_folder, "sw.js",
                               mimetype="application/javascript")


@app.errorhandler(Exception)
def handle_exception(e):
    import traceback
    err_tb = traceback.format_exc()
    print("=== UNHANDLED EXCEPTION ===", file=sys.stderr)
    print(err_tb, file=sys.stderr)
    print("===========================", file=sys.stderr)
    # If debug mode or explicitly caught 500
    return f"<h2>Internal Server Error</h2><p style='color:red;'><b>{type(e).__name__}:</b> {e}</p><pre style='background:#f1f5f9;padding:12px;border-radius:6px;overflow-x:auto;'>{err_tb}</pre>", 500


# ---------------------------------------------------------------------------
# Routes — Authentication & Dashboard
# ---------------------------------------------------------------------------
@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        db  = get_db()
        row = db.execute(
            text("SELECT * FROM users WHERE username = :u"), {"u": username}
        ).mappings().fetchone()
        if row and check_password_hash(row["password_hash"], password):
            session["username"] = row["username"]
            session["role"]     = row["role"]
            return redirect(url_for("dashboard"))
        flash("Invalid username or password.", "error")
    return render_template("login.html")


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.route("/")
@login_required
def dashboard():
    db    = get_db()
    items = db.execute(
        text(f"SELECT * FROM inventory ORDER BY {ORDER_BY}")
    ).mappings().fetchall()
    
    users = []
    if session.get("role") == "admin":
        users = db.execute(
            text("SELECT username, role FROM users ORDER BY username")
        ).mappings().fetchall()

    return render_template(
        "index.html",
        items=items,
        users=users,
        role=session.get("role"),
        username=session.get("username"),
    )


# ---------------------------------------------------------------------------
# Routes — Inventory Management
# ---------------------------------------------------------------------------
@app.route("/add", methods=["POST"])
@login_required
@admin_required
def add_item():
    material = request.form.get("material", "").strip()
    try:
        buying    = float(request.form.get("buying_price", 0))
        wholesale = float(request.form.get("wholesale_price", 0))
        retail    = float(request.form.get("retail_price", 0))
    except ValueError:
        flash("Prices must be valid numbers.", "error")
        return redirect(url_for("dashboard"))

    if not material:
        flash("Material name is required.", "error")
        return redirect(url_for("dashboard"))

    db = get_db()
    db.execute(
        text("""INSERT INTO inventory
                 (material, buying_price, wholesale_price, retail_price, updated_at)
                 VALUES (:m, :b, :w, :r, :u)"""),
        {"m": material, "b": buying, "w": wholesale, "r": retail,
         "u": datetime.now().strftime("%Y-%m-%d %H:%M")},
    )
    db.commit()
    flash(f"Added item '{material}'.", "success")
    return redirect(url_for("dashboard"))


@app.route("/edit/<int:item_id>", methods=["POST"])
@login_required
@admin_required
def edit_item(item_id):
    material = request.form.get("material", "").strip()
    try:
        buying    = float(request.form.get("buying_price", 0))
        wholesale = float(request.form.get("wholesale_price", 0))
        retail    = float(request.form.get("retail_price", 0))
    except ValueError:
        flash("Prices must be valid numbers.", "error")
        return redirect(url_for("dashboard"))

    db = get_db()
    db.execute(
        text("""UPDATE inventory
                SET material=:m, buying_price=:b, wholesale_price=:w,
                    retail_price=:r, updated_at=:u
                WHERE id=:id"""),
        {"m": material, "b": buying, "w": wholesale, "r": retail,
         "u": datetime.now().strftime("%Y-%m-%d %H:%M"), "id": item_id},
    )
    db.commit()
    flash(f"Updated item '{material}'.", "success")
    return redirect(url_for("dashboard"))


@app.route("/delete/<int:item_id>", methods=["POST"])
@login_required
@admin_required
def delete_item(item_id):
    db  = get_db()
    row = db.execute(
        text("SELECT material FROM inventory WHERE id = :id"), {"id": item_id}
    ).mappings().fetchone()
    db.execute(text("DELETE FROM vendor_payments WHERE inventory_id = :id"), {"id": item_id})
    db.execute(text("DELETE FROM inventory WHERE id = :id"), {"id": item_id})
    db.commit()
    if row:
        flash(f"Deleted item '{row['material']}'.", "success")
    return redirect(url_for("dashboard"))


# ---------------------------------------------------------------------------
# Routes — Vendor Details
# ---------------------------------------------------------------------------
@app.route("/vendor/<int:item_id>", methods=["GET"])
@login_required
@admin_required
def vendor_get(item_id):
    from flask import jsonify
    db  = get_db()
    row = db.execute(
        text("""SELECT v.*, i.updated_at
                FROM inventory i
                LEFT JOIN vendor_payments v ON i.id = v.inventory_id
                WHERE i.id = :id"""),
        {"id": item_id}
    ).mappings().fetchone()
    if row:
        data = dict(row)
        if data.get("inventory_id") is None:
            data["inventory_id"] = item_id
            data["vendor_name"]  = ""
            data["vendor_phone"] = ""
            data["total_amount"] = 0
            data["paid_amount"]  = 0
    else:
        data = {"inventory_id": item_id, "vendor_name": "", "vendor_phone": "",
                "total_amount": 0, "paid_amount": 0, "updated_at": ""}
    data["left_to_pay"] = round((data.get("total_amount") or 0) - (data.get("paid_amount") or 0), 2)
    return jsonify(data)


@app.route("/vendor/<int:item_id>", methods=["POST"])
@login_required
@admin_required
def vendor_save(item_id):
    from flask import jsonify
    payload      = request.get_json(force=True) or {}
    vendor_name  = str(payload.get("vendor_name",  "")).strip()
    vendor_phone = str(payload.get("vendor_phone", "")).strip()
    try:
        total_amount = float(payload.get("total_amount", 0))
        paid_amount  = float(payload.get("paid_amount",  0))
    except (ValueError, TypeError):
        return jsonify({"error": "Amounts must be numbers."}), 400

    db  = get_db()
    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    existing = db.execute(
        text("SELECT inventory_id FROM vendor_payments WHERE inventory_id = :id"), {"id": item_id}
    ).fetchone()
    if existing:
        db.execute(
            text("""UPDATE vendor_payments
                     SET vendor_name=:vn, vendor_phone=:vp,
                         total_amount=:ta, paid_amount=:pa, updated_at=:ua
                     WHERE inventory_id=:id"""),
            {"vn": vendor_name, "vp": vendor_phone, "ta": total_amount,
             "pa": paid_amount, "ua": now, "id": item_id},
        )
    else:
        db.execute(
            text("""INSERT INTO vendor_payments
                     (inventory_id, vendor_name, vendor_phone, total_amount, paid_amount, updated_at)
                     VALUES (:id, :vn, :vp, :ta, :pa, :ua)"""),
            {"id": item_id, "vn": vendor_name, "vp": vendor_phone,
             "ta": total_amount, "pa": paid_amount, "ua": now},
        )
    db.commit()
    return jsonify({"ok": True, "left_to_pay": round(total_amount - paid_amount, 2)})


# ---------------------------------------------------------------------------
# Routes — Excel Export (admin only)
# ---------------------------------------------------------------------------
@app.route("/export")
@login_required
@admin_required
def export_excel():
    import io
    from flask import send_file
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    db    = get_db()
    items = db.execute(
        text(f"SELECT * FROM inventory ORDER BY {ORDER_BY}")
    ).mappings().fetchall()
    vendors = db.execute(
        text("SELECT * FROM vendor_payments")
    ).mappings().fetchall()
    vendor_map = {v["inventory_id"]: v for v in vendors}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventory"

    # Styles
    header_font    = Font(bold=True, color="FFFFFF", size=11)
    header_fill    = PatternFill("solid", fgColor="1F2D3D")
    subhdr_fill    = PatternFill("solid", fgColor="2E4054")
    subhdr_font    = Font(bold=True, color="FFFFFF", size=10)
    center         = Alignment(horizontal="center", vertical="center")
    right          = Alignment(horizontal="right",  vertical="center")
    thin           = Side(style="thin", color="D2D5C9")
    border         = Border(left=thin, right=thin, top=thin, bottom=thin)
    amber_fill     = PatternFill("solid", fgColor="E2952E")
    amber_font     = Font(bold=True, color="2B1D06", size=10)

    # Title row
    ws.merge_cells("A1:I1")
    title_cell = ws["A1"]
    title_cell.value     = "Banik Hardware — Inventory Report"
    title_cell.font      = Font(bold=True, size=14, color="1F2D3D")
    title_cell.alignment = center
    ws.row_dimensions[1].height = 28

    # Subtitle
    ws.merge_cells("A2:I2")
    sub_cell = ws["A2"]
    sub_cell.value     = f"Exported: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    sub_cell.font      = Font(italic=True, size=9, color="6B7A8F")
    sub_cell.alignment = center
    ws.row_dimensions[2].height = 16

    # Group headers row 3
    ws.merge_cells("A3:A3"); ws["A3"].value = "#"
    ws.merge_cells("B3:B3"); ws["B3"].value = "Material"
    ws.merge_cells("C3:E3"); ws["C3"].value = "Pricing"
    ws.merge_cells("F3:F3"); ws["F3"].value = "Last Modified"
    ws.merge_cells("G3:I3"); ws["G3"].value = "Vendor & Payments"
    for col in ["A3","B3","C3","F3","G3"]:
        ws[col].font      = header_font
        ws[col].fill      = header_fill
        ws[col].alignment = center
        ws[col].border    = border
    ws.row_dimensions[3].height = 20

    # Sub-headers row 4
    sub_hdrs = ["#", "Material", "Buying", "Wholesale", "Retail",
                "Modified", "Vendor Name", "Phone", "Total", "Paid", "Left to Pay"]
    # We have 11 columns: A-K
    # Adjust group headers
    ws.merge_cells("C3:E3"); ws["C3"].value = "Pricing"
    ws.merge_cells("G3:K3"); ws["G3"].value = "Vendor & Payments"
    for cell in ["A3","B3","C3","F3","G3"]:
        ws[cell].font      = header_font
        ws[cell].fill      = header_fill
        ws[cell].alignment = center
        ws[cell].border    = border

    cols = ["A","B","C","D","E","F","G","H","I","J","K"]
    sub_labels = ["#","Material","Buying","Wholesale","Retail",
                  "Modified","Vendor Name","Phone","Total Amt","Paid Amt","Left to Pay"]
    for i, (col, lbl) in enumerate(zip(cols, sub_labels)):
        cell = ws[f"{col}4"]
        cell.value     = lbl
        cell.font      = subhdr_font
        cell.fill      = subhdr_fill
        cell.alignment = center
        cell.border    = border
    ws.row_dimensions[4].height = 18

    # Data rows starting row 5
    num_fmt = '#,##0.00'
    for row_i, item in enumerate(items, start=1):
        r    = row_i + 4
        vend = vendor_map.get(item["id"], {})
        total = float(vend.get("total_amount", 0) or 0)
        paid  = float(vend.get("paid_amount",  0) or 0)
        left  = round(total - paid, 2)

        row_data = [
            row_i,
            item["material"],
            item["buying_price"],
            item["wholesale_price"],
            item["retail_price"],
            item["updated_at"],
            vend.get("vendor_name",  "") or "",
            vend.get("vendor_phone", "") or "",
            total,
            paid,
            left,
        ]
        fill_bg = PatternFill("solid", fgColor="F9F8F4") if row_i % 2 == 0 else PatternFill()
        for col_i, (col, val) in enumerate(zip(cols, row_data)):
            cell = ws[f"{col}{r}"]
            cell.value  = val
            cell.border = border
            if fill_bg.fill_type:
                cell.fill = fill_bg
            if col_i in (2, 3, 4, 8, 9):   # price columns
                cell.number_format = num_fmt
                cell.alignment     = right
            elif col_i == 10:               # left to pay — highlight if > 0
                cell.number_format = num_fmt
                cell.alignment     = right
                if left > 0:
                    cell.font = Font(bold=True, color="B4432F")
            elif col_i == 0:                # row number
                cell.alignment = center
        ws.row_dimensions[r].height = 16

    # Column widths
    col_widths = [5, 28, 13, 13, 13, 16, 22, 15, 13, 13, 13]
    for col, w in zip(cols, col_widths):
        ws.column_dimensions[col].width = w

    # Freeze panes below sub-headers
    ws.freeze_panes = "A5"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"inventory_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ---------------------------------------------------------------------------
# Routes — User Account Management
# ---------------------------------------------------------------------------
@app.route("/users/create", methods=["POST"])
@login_required
@admin_required
def create_user():
    username = request.form.get("username", "").strip()
    password = request.form.get("password", "")
    role     = request.form.get("role", "user").strip()

    if not username or not password:
        flash("Username and password are required.", "error")
        return redirect(url_for("dashboard"))

    if role not in ("admin", "user"):
        role = "user"

    db = get_db()
    existing = db.execute(
        text("SELECT username FROM users WHERE username = :u"), {"u": username}
    ).fetchone()
    if existing:
        flash(f"User '{username}' already exists.", "error")
        return redirect(url_for("dashboard"))

    db.execute(
        text("INSERT INTO users (username, password_hash, role) VALUES (:u, :p, :r)"),
        {"u": username, "p": generate_password_hash(password), "r": role}
    )
    db.commit()
    flash(f"User '{username}' created successfully as {role.upper()}.", "success")
    return redirect(url_for("dashboard"))


@app.route("/users/change-password", methods=["POST"])
@login_required
def change_password():
    target_username = request.form.get("username", "").strip()
    new_password    = request.form.get("new_password", "")

    if not new_password:
        flash("New password cannot be empty.", "error")
        return redirect(url_for("dashboard"))

    # Non-admin users can only change their own password
    if session.get("role") != "admin":
        target_username = session.get("username")
    elif not target_username:
        target_username = session.get("username")

    db = get_db()
    user = db.execute(
        text("SELECT username FROM users WHERE username = :u"), {"u": target_username}
    ).fetchone()

    if not user:
        flash("User not found.", "error")
        return redirect(url_for("dashboard"))

    db.execute(
        text("UPDATE users SET password_hash = :p WHERE username = :u"),
        {"p": generate_password_hash(new_password), "u": target_username}
    )
    db.commit()
    flash(f"Password updated for '{target_username}'.", "success")
    return redirect(url_for("dashboard"))


@app.route("/users/delete/<path:username>", methods=["POST"])
@login_required
@admin_required
def delete_user(username):
    if username == session.get("username"):
        flash("You cannot delete your own active account.", "error")
        return redirect(url_for("dashboard"))

    db = get_db()
    db.execute(text("DELETE FROM users WHERE username = :u"), {"u": username})
    db.commit()
    flash(f"User account '{username}' removed.", "success")
    return redirect(url_for("dashboard"))


# ---------------------------------------------------------------------------
# Local IP helper
# ---------------------------------------------------------------------------
def get_local_ip():
    s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
    try:
        s.connect(("8.8.8.8", 80))
        ip = s.getsockname()[0]
    except Exception:
        ip = "127.0.0.1"
    finally:
        s.close()
    return ip


if __name__ == "__main__":
    import threading
    import webbrowser

    ip = get_local_ip()
    print("=" * 60)
    print("  Banik Hardware — Inventory is running!")
    print(f"  This PC:        http://127.0.0.1:5000")
    print(f"  Other devices:  http://{ip}:5000")
    print("=" * 60)

    threading.Timer(1.5, lambda: webbrowser.open("http://127.0.0.1:5000")).start()
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 5000)), debug=False)
